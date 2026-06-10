import datetime
from openpyxl import load_workbook

# Use raw strings for Windows paths
path = r'D:\Project\Wischos trading website\二次工作文件\Wischos_客户开发CRM.xlsx'
wb = load_workbook(path)
today = datetime.datetime(2026, 5, 30)
new_id = 'C003'

# 1. Update '客户库'
ws_lib = wb['客户库']
row_lib = 4 # Start checking from row 4 (after header)
while ws_lib.cell(row=row_lib, column=1).value:
    row_lib += 1

ws_lib.cell(row=row_lib, column=1, value=new_id)
ws_lib.cell(row=row_lib, column=2, value='Brandz Marketing Inc.')
ws_lib.cell(row=row_lib, column=3, value='Graeme Dawes')
ws_lib.cell(row=row_lib, column=4, value='CEO & President')
ws_lib.cell(row=row_lib, column=5, value='分销商')
ws_lib.cell(row=row_lib, column=6, value='https://brandzmarketing.com/')
ws_lib.cell(row=row_lib, column=10, value='Stony Plain, AB · 加拿大')

# 2. Update '主板'
ws_main = wb['主板'] if '主板' in wb.sheetnames else wb.active
row_main = 7
while ws_main.cell(row=row_main, column=1).value:
    row_main += 1

ws_main.cell(row=row_main, column=1, value=new_id)
# Formula for Hyperlink and VLOOKUP
formula_name = f'=IFERROR(HYPERLINK("#客户库!A"&MATCH(A{row_main},客户库!$A:$A,0),VLOOKUP(A{row_main},客户库!$A:$B,2,0)),A{row_main})'
ws_main.cell(row=row_main, column=2, value=formula_name)
ws_main.cell(row=row_main, column=3, value='Graeme Dawes')
ws_main.cell(row=row_main, column=4, value='分销商')
ws_main.cell(row=row_main, column=5, value='加拿大')
ws_main.cell(row=row_main, column=6, value='已发消息')
ws_main.cell(row=row_main, column=7, value='A')
ws_main.cell(row=row_main, column=8, value=today)
# Formula for follow-up count
formula_count = f'=IFERROR(COUNTIF(跟进日志!$C:$C,A{row_main}),0)'
ws_main.cell(row=row_main, column=10, value=formula_count)

# 3. Update '跟进日志'
ws_log = wb['跟进日志']
row_log = 3
while ws_log.cell(row=row_log, column=3).value:
    row_log += 1

# ID Auto-increment formula
formula_seq = f'=IF(C{row_log}="","",COUNTA($C$3:C{row_log}))'
ws_log.cell(row=row_log, column=1, value=formula_seq)
ws_log.cell(row=row_log, column=2, value=today)
ws_log.cell(row=row_log, column=3, value=new_id)
ws_log.cell(row=row_log, column=4, value='Brandz Marketing Inc.')
ws_log.cell(row=row_log, column=5, value='Graeme Dawes')
ws_log.cell(row=row_log, column=6, value='LinkedIn')
ws_log.cell(row=row_log, column=7, value='出')
ws_log.cell(row=row_log, column=8, value='发送LinkedIn连接邀请：Hi Graeme, really respect the 30-year legacy you built in the Tri-Region! I noticed your collections are strong but missing high-end metal gift sets. We craft premium metal kits in China that could fill that gap. Connect?')
ws_log.cell(row=row_log, column=9, value='待通过')
ws_log.cell(row=row_log, column=10, value='待接触 → 已发消息')

wb.save(path)
print("CRM updated successfully.")
