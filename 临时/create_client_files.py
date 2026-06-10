#!/usr/bin/env python3
"""
Create client Excel files for Branding Edge and Great Gift Ideas,
move their MD files to proper Canada subfolders, and update CRM.
"""
import os
import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = "/mnt/d/Project/Wischos trading website/临时"

# ── Helpers ──────────────────────────────────────────────────────────────────

def hdr(ws, text, row=1):
    """Bold header cell A{row}."""
    ws.cell(row=row, column=1, value=text).font = Font(bold=True, size=12)

def subhdr(ws, text, row=2):
    ws.cell(row=row, column=1, value=text).font = Font(italic=True, color="595959")

def section(ws, label, row):
    """Grey section divider across columns A-B."""
    c = ws.cell(row=row, column=1, value=label)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="595959")
    ws.merge_cells(f"A{row}:B{row}")

def kv(ws, key, value, row):
    ws.cell(row=row, column=1, value=key).font = Font(bold=True)
    ws.cell(row=row, column=2, value=value)

def th(ws, headers, row):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2F5597")

def row_data(ws, values, r):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.alignment = Alignment(wrap_text=True)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ══════════════════════════════════════════════════════════════════════════════
# BRANDING EDGE – Graham Ellerman  (C038)
# ══════════════════════════════════════════════════════════════════════════════

def make_branding_edge(path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1 ───────────────────────────────────────────────────────────────
    ws = wb.create_sheet("1-客户档案")
    hdr(ws, "Branding Edge — 客户档案")
    subhdr(ws, "建档日期：2026-05-14    来源：官网 / LinkedIn / ZoomInfo")
    r = 3
    kv(ws, "项目", "内容", r); ws["A3"].font = Font(bold=True); ws["B3"].font = Font(bold=True); r+=1
    section(ws, "【基础信息】", r); r+=1
    for k, v in [
        ("客户ID", "C038"),
        ("公司名", "Branding Edge"),
        ("联系人", "Graham Ellerman"),
        ("职位", "Business Owner"),
        ("客户类型", "小型本地促销品经销商 / 品牌定制装饰商"),
        ("地区", "Devon / Edmonton, Alberta, Canada"),
        ("网站", "https://brandingedge.ca"),
        ("邮箱/电话", "g***@brandingedge.ca（待确认）"),
    ]:
        kv(ws, k, v, r); r+=1
    section(ws, "【一轮判断】", r); r+=1
    for k, v in [
        ("综合优先级", "★★★（3/5）— B边缘"),
        ("当前阶段", "已发消息"),
        ("Fit Score", "55/100"),
        ("Value Score", "54/100"),
        ("开发投入级别", "Light-touch（LinkedIn邀请+一封邮件+一次跟进；无回复30天暂停）"),
        ("核心判断", "小型本地经销商；有内部激光雕刻机；油气/工业客户群；现有奖品停留在树脂/亚克力层级，高端金属空白。"),
        ("首轮限制", "不推金属奖牌/纪念币/lapel pins（非Wischos产品线）。推黄铜笔/桌面套装/保温杯。"),
    ]:
        kv(ws, k, v, r); r+=1
    section(ws, "【供应商背景】", r); r+=1
    for k, v in [
        ("现有供应商", "Koozie Group（BIC）/ Starline / Bullet（北美本地目录采购，无直接中国进货证据）"),
        ("个性化技术能力", "内部激光雕刻机（in-house）—可在金属件上刻logo，自行增值"),
        ("主要产品线", "FR阻燃工装 / 企业制服 / 运动套装 / 奖品/雕刻 / 促销礼品 / 印刷标识 / 3D打印"),
        ("信息来源", "ZoomInfo（Graham/Cass档案）+ 官网 + LinkedIn（2026-05-14）"),
    ]:
        kv(ws, k, v, r); r+=1
    set_col_widths(ws, [22, 70])

    # ── Sheet 2 ───────────────────────────────────────────────────────────────
    ws = wb.create_sheet("2-联系人")
    hdr(ws, "联系人档案")
    r = 2
    th(ws, ["姓名","职位","决策权","经验背景","LinkedIn","Email","电话","首选沟通渠道","备注"], r); r+=1
    row_data(ws, [
        "Graham Ellerman","Business Owner","运营/生产/客户交付；主要决策人",
        "NAIT锅炉工学位；前Aptim工头（石油工程服务）；实用主义/工业背景",
        "LinkedIn 3度好友","g***@brandingedge.ca","—","LinkedIn → Email",
        "主触达对象；不喜欢销售废话；重视产品能否用/交期能否到"
    ], r); r+=1
    row_data(ws, [
        "Cass Ellerman","Company Owner","采购审批/财务决策",
        "MacEwan大学商务管理；前Gibson Energy业务分析师",
        "—","（同域名）","—","邮件引出",
        "先从Graham进；邮件来往后自然引出Cass是否参与"
    ], r)
    set_col_widths(ws, [18,25,20,35,30,28,14,18,40])
    for row in ws.iter_rows(min_row=3, max_row=4):
        for c in row:
            c.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[3].height = 50
    ws.row_dimensions[4].height = 50

    # ── Sheet 3 ───────────────────────────────────────────────────────────────
    ws = wb.create_sheet("3-切入角度")
    hdr(ws, "三个切入角度（按优先级排序）")
    r = 2
    th(ws, ["排序","角度","为什么有效","话术 Hook","风险/注意"], r); r+=1
    angles = [
        ("⭐⭐⭐⭐", "为激光雕刻机配高档金属坯件",
         "已有内部雕刻机但刻树脂/亚克力；金属件成品档次跳升；可卖更高价",
         "\"Saw you do in-house laser engraving — we supply brass pens, metal card cases and tumblers that engrave cleanly and lift perceived value.\"",
         "可能已有便宜金属坯件来源"),
        ("⭐⭐⭐", "油气/建筑客户安全表彰+承包商礼品",
         "FR工装客户（油气/建筑）→ 固定安全里程碑奖励+年终承包商感谢礼品预算；金属套装比塑料礼品档次高",
         "\"Do your oil & gas or construction clients do safety milestone recognition or year-end appreciation gifts? Our metal gift sets go well beyond the usual acrylic plaque.\"",
         "MOQ 100套需客户项目背书"),
        ("⭐⭐", "项目型备用供应商（收尾留门）",
         "无立即需求时保持连接",
         "\"If you ever get a large year-end or recognition brief that needs premium metal — we can support the sourcing.\"",
         "纯被动，转化率低"),
    ]
    for a in angles:
        row_data(ws, a, r); r+=1
    set_col_widths(ws, [8,28,38,55,28])

    # ── Sheet 4 ───────────────────────────────────────────────────────────────
    ws = wb.create_sheet("4-推荐产品配对")
    hdr(ws, "场景 × Wischos 产品配对")
    r = 2
    th(ws, ["场景","推荐 SKU","产品名","为什么适合","Brand侧卖法/客户场景"], r); r+=1
    products = [
        ("工业/建筑客户安全里程碑认可","WP-101 / WP-102","Brass Crown Bolt-Action / Brass Rollerball",
         "黄铜笔有重量感；比塑料笔档次高10倍；可刻名字/里程碑日期",
         "安全里程碑奖励（100/500/1000天无事故）"),
        ("桌面礼品套装（VIP/承包商）","WGS-005","Executive Desk Set",
         "多件组合，呈现感强；适合年终VIP承包商礼品",
         "年终承包商感谢礼品；含包装，$120–180/套"),
        ("员工表彰/奖励礼盒","WGS-006","The Onboarding Set",
         "适合员工认可/奖励场景；全铝套件",
         "员工里程碑/表彰礼盒"),
        ("激光雕刻配件——黄铜坯件","WP-101 / WP-203 / WP-401","笔/信封刀/保温杯",
         "内部激光雕刻机可直接雕刻；比树脂/亚克力档次高；增值利润更大",
         "配合Branding Edge内部雕刻，成品更高档"),
    ]
    for p in products:
        row_data(ws, p, r); r+=1
    set_col_widths(ws, [28,18,28,40,35])

    # ── Sheet 5 ───────────────────────────────────────────────────────────────
    ws = wb.create_sheet("5-跟进时间表")
    hdr(ws, "多触点跟进策略（Light-touch）")
    r = 2
    th(ws, ["Day","渠道","动作","话术/物料","状态","实际日期/备注"], r); r+=1
    timeline = [
        ("Day 0","LinkedIn","发送Graham邀请",
         "\"Like how Branding Edge covers the full industrial recognition journey — FR gear on the job, in-house engraving for the milestone moments. I help promo dealers build custom metal gift options from China — brass, titanium, stainless pieces. Thought it may be worth connecting.\"",
         "✅ 已发送","2026-05-14 LinkedIn邀请已发（版本B，无近期发帖）"),
        ("连接后立即","LinkedIn DM","发Touch 0私信（建立认知）",
         "Hi Graham, thanks for connecting. Just to give a clearer picture of what I do...",
         "⬜ 待执行","连接通过后立即发"),
        ("Day 7","Email","Touch 1首封邮件（一个问题，不发目录）",
         "Subject: Metal blanks for your engraving clients — Branding Edge",
         "⬜ 待执行",""),
        ("Day 10–12","Email","Touch 2产品规格聚焦",
         "Subject: Re: Metal blanks for your engraving clients",
         "⬜ 待执行",""),
        ("Day 14–16","Email","Touch 3行业洞察",
         "Subject: Recognition gifting in oil & gas — a quick note",
         "⬜ 待执行",""),
        ("Day 18–21","Email","Touch 4附WGS-005图片",
         "Subject: One photo — unbranded desk set",
         "⬜ 待执行",""),
        ("Day 28–30","Email","Touch 5收尾留门",
         "Subject: Closing the loop — Branding Edge",
         "⬜ 待执行",""),
        ("Day 30","CRM","无回复则暂停","—","待触发",""),
        ("Day 90","LinkedIn/网站","重查触发信号（新帖/展会/网站更新）","—","待触发",""),
    ]
    for t in timeline:
        row_data(ws, t, r); r+=1
    set_col_widths(ws, [14,14,30,60,14,30])

    # ── Sheet 6 ───────────────────────────────────────────────────────────────
    ws = wb.create_sheet("6-话术库")
    hdr(ws, "话术库（LinkedIn / DM / Email / 异议预案）")
    r = 2
    th(ws, ["类型","内容","备注/设计理由"], r); r+=1
    scripts = [
        ("LinkedIn邀请（✅ 已发送 2026-05-14）",
         "Like how Branding Edge covers the full industrial recognition journey — FR gear on the job, in-house engraving for the milestone moments. I help promo dealers build custom metal gift options from China — brass, titanium, stainless pieces. Thought it may be worth connecting.",
         "版本B（无近期发帖）；~265 chars；强调FR工装+激光雕刻两端"),
        ("Touch 0 — 连接后私信",
         "Hi Graham, thanks for connecting. Just to give a clearer picture of what I do: I work mainly with metal gift projects — brass, titanium and stainless pieces for corporate gifting, recognition and desk-use programs.\n\nI'm not limited to catalogue items; most pieces can be built around a client brief, material or finish. I'll share relevant ideas occasionally — if a brief involving custom metal comes up, happy to help explore options.\n\nThanks,\n[姓名 / Wischos Gift]",
         "建立认知；不重复邀请措辞；不发目录"),
        ("Touch 1 — 首封邮件 (Day 7)",
         "Subject: Metal blanks for your engraving clients — Branding Edge\n\nHi Graham,\n\nNoticed Branding Edge does in-house laser engraving and serves industrial clients in the Edmonton area.\n\nQuick question: do your oil & gas or construction clients ever need safety milestone awards or year-end appreciation gifts that go beyond the usual acrylic plaque — something with actual weight to it?\n\nI source custom metal gift pieces from China — brass pens, desk accessories, presentation sets — for promo dealers whose engravers can personalise them in-house.\n\nHappy to share a photo if that's useful.\n\n[姓名]\nWischos Gift | [邮箱] | wischosgift.com\nTo opt out of future messages, reply with \"unsubscribe\".",
         "研究型开场；一个问题；不发目录；含CASL退订"),
        ("Touch 2 — Day 10–12",
         "Subject: Re: Metal blanks for your engraving clients\n\nHi Graham,\n\nFollowing up briefly. The pieces that tend to work well for industrial recognition programs:\n- Solid brass pens — lacquered or bare brass, both engrave cleanly\n- Metal card cases in stainless or aluminium\n- Desk sets: pen + card case + stand, boxed presentation\n\nThese sit at a price point that feels substantial as a milestone gift — better retention than plastic promo, but not so expensive they need a long approval process.\n\nDo specs like these match what your clients typically ask for in recognition programs?\n\n[姓名]",
         "产品规格聚焦；问匹配度"),
        ("Touch 3 — Day 14–16",
         "Subject: Recognition gifting in oil & gas — a quick note\n\nHi Graham,\n\nA quick note on context: recognition gifting in oil & gas and construction usually falls into two buckets — safety milestones (100/500/1,000 days without incident) and annual contractor or vendor appreciation programs.\n\nBoth typically need something that can be personalised with a name or logo and looks credible on a desk. Most of that market still uses acrylic plaques. Metal pieces at a similar price point tend to be better received and kept longer.\n\nDo you currently source anything like that for your industrial clients?\n\n[姓名]",
         "行业洞察；建立专业感"),
        ("Touch 4 — Day 18–21（附图）",
         "Subject: One photo — unbranded desk set\n\nHi Graham,\n\nRather than describe further — attaching a photo of one of our unbranded desk sets [附WGS-005图]. Brass pen, metal card case, presentation box.\n\nYour engraver could add the client's name or logo directly. I'd handle production and ship to you or direct to your client.\n\nIf nothing fits right now, that's completely fine — just wanted to put something concrete in front of you.\n\n[姓名]",
         "附WGS-005图；低门槛；不强推"),
        ("Touch 5 — Day 28–30（收尾）",
         "Subject: Closing the loop — Branding Edge\n\nHi Graham,\n\nI'll keep this short — if custom metal pieces for recognition or corporate gifting programs aren't relevant to Branding Edge right now, no problem at all.\n\nIf there's a better person at your end to connect with on sourcing, happy to be pointed in the right direction.\n\nOtherwise I'll leave it here for now and check back if something relevant comes up.\n\n[姓名]",
         "收尾留门；低压力"),
        ("异议：已有供应商",
         "品类扩展不是替换——我补你目录里没有的高档金属层级。有项目需要时可作备用选项。",
         ""),
        ("异议：MOQ太高",
         "标准100套；有具体客户项目时可先了解需求再讨论。",
         ""),
        ("异议：海外交期太长",
         "生产通常25–35天。告诉我时间节点，我评估能否配合。",
         ""),
        ("异议：只用本地供应",
         "尊重。记入CRM作市场情报，90天后重查。",
         ""),
    ]
    for s in scripts:
        row_data(ws, s, r); r+=1
    set_col_widths(ws, [28,80,35])

    # ── Sheet 7 ───────────────────────────────────────────────────────────────
    ws = wb.create_sheet("7-Sample-Pack")
    hdr(ws, "Sample Pack（Branding Edge 轻量版）")
    ws.cell(row=2, column=1, value="策略：连接通过并有兴趣后准备；先发产品图，不主动寄样。确认项目需求后再寄。")
    r = 3
    th(ws, ["#","SKU","产品","定制内容","用途/演示"], r); r+=1
    samples = [
        (1,"WP-101","Brass Crown Bolt-Action Pen","logo/name engraving","激光雕刻演示——证明金属件可直接用内部雕刻机处理"),
        (2,"WP-203","Brass Letter Opener","logo engraving","桌面recognition件；比亚克力奖品档次高"),
        (3,"WGS-005","Executive Desk Set（图片）","—（先发图）","向Graham展示完整套装呈现感；Touch 4附图"),
        (4,"WP-401","Pure Titanium Bottle（规格单）","logo engraving","油气/建筑高端礼品；材质差异化"),
    ]
    for s in samples:
        row_data(ws, s, r); r+=1
    set_col_widths(ws, [4,18,30,25,45])

    # ── Sheet 8 ───────────────────────────────────────────────────────────────
    ws = wb.create_sheet("8-行动清单")
    hdr(ws, "本周可执行清单 + 长期扩展")
    r = 2
    th(ws, ["#","任务","类型","状态","完成日期/备注"], r); r+=1
    actions = [
        (1,"二轮客户调研完成（见MD文件）","内部研究","✅ 已完成","2026-05-14"),
        (2,"LinkedIn邀请已发（版本B）","外联","✅ 已完成","2026-05-14"),
        (3,"查Graham近30天LinkedIn发帖","外联准备","⬜ 待执行","若有发帖则选版本A话术"),
        (4,"建立客户文件夹 + 创建本Excel","内部整理","✅ 已完成","2026-05-14"),
        (5,"同步写入总CRM：C038","内部整理","⬜ 待执行",""),
        (6,"连接通过后发Touch 0私信","Day 0+","⬜ 待触发",""),
        (7,"Day 7 发Touch 1首封邮件","Day 7","⬜ 待触发",""),
        (8,"Day 10–12 发Touch 2","Day 10–12","⬜ 待触发",""),
        (9,"Day 14–16 发Touch 3","Day 14–16","⬜ 待触发",""),
        (10,"Day 18–21 发Touch 4（附WGS-005图）","Day 18–21","⬜ 待触发",""),
        (11,"Day 28–30 发Touch 5收尾","Day 28–30","⬜ 待触发",""),
        (12,"Day 30 无回复 → 暂停，CRM标注","Day 30","⬜ 待触发",""),
        (13,"Day 90 重查LinkedIn/网站触发信号","Day 90","⬜ 待触发",""),
    ]
    for a in actions:
        row_data(ws, a, r); r+=1
    set_col_widths(ws, [4,45,18,14,30])

    # ── Sheet 9 ───────────────────────────────────────────────────────────────
    ws = wb.create_sheet("9-备注与扩展")
    hdr(ws, "长期扩展空间 + 关键备注")
    r = 2
    section(ws, "【长期扩展】", r); r+=1
    notes = [
        ("首单成后","探索年度油气/建筑客户安全里程碑固定采购计划"),
        ("Cass引入路径","Graham确认兴趣后 → 自然引出Cass参与采购审批"),
        ("品类扩展","若首单顺利，推保温杯/饮品品类，扩大单均价"),
    ]
    for k, v in notes:
        kv(ws, k, v, r); r+=1
    section(ws, "【关键提醒】", r); r+=1
    reminders = [
        ("产品定位纠正","不推金属奖牌/纪念币/lapel pins/challenge coins——这些不是Wischos产品线"),
        ("正确切入品","黄铜笔（WP-101/102）/ 高管桌面套装（WGS-005）/ 奖励套装（WGS-006）/ 保温杯 / 金属名片夹"),
        ("供应商切换逻辑","品类扩展（非替换现有供应商）；Koozie/Bullet/Starline目录里没有精品金属礼品套装"),
        ("切换阻力","习惯本地目录采购；25–35天周期可能冲突；小公司难承诺100套MOQ"),
        ("CASL合规","官网公开展示商业联系方式；冷邮件须含：发件人身份+公司联系信息+退订机制"),
        ("体量天花板","预估年1–3单；轻度主动跟进；不过度投入"),
    ]
    for k, v in reminders:
        kv(ws, k, v, r); r+=1
    section(ws, "【数据来源】", r); r+=1
    sources = [
        ("ZoomInfo – Graham","https://www.zoominfo.com/p/Graham-Ellerman/8556358931"),
        ("ZoomInfo – Cass","https://www.zoominfo.com/p/Cass-Ellerman/8494899848"),
        ("官网","https://brandingedge.ca/"),
    ]
    for k, v in sources:
        kv(ws, k, v, r); r+=1
    set_col_widths(ws, [22,70])

    wb.save(path)
    print(f"✓ Saved: {path}")


# ══════════════════════════════════════════════════════════════════════════════
# GREAT GIFT IDEAS – Lisa Janssen  (C039)
# ══════════════════════════════════════════════════════════════════════════════

def make_great_gift_ideas(path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1 ───────────────────────────────────────────────────────────────
    ws = wb.create_sheet("1-客户档案")
    hdr(ws, "Great Gift Ideas — 客户档案")
    subhdr(ws, "建档日期：2026-05-14    来源：官网 / LinkedIn / ASI会员验证")
    r = 3
    kv(ws, "项目", "内容", r); ws["A3"].font = Font(bold=True); ws["B3"].font = Font(bold=True); r+=1
    section(ws, "【基础信息】", r); r+=1
    for k, v in [
        ("客户ID", "C039"),
        ("公司名", "Great Gift Ideas"),
        ("联系人", "Lisa Janssen"),
        ("职位", "Owner"),
        ("客户类型", "Boutique promotional products distributor / corporate gifting agency"),
        ("地区", "Richmond Hill, Ontario, Canada"),
        ("网站", "greatgiftideas.ca"),
        ("邮箱/电话", "lisa@greatgiftideas.ca（官网确认） / (647) 224-2399"),
        ("成立年份", "2003（Lisa独立经营20+年）"),
        ("行业协会", "ASI会员（Powered by ASI，3,000+供应商网络）"),
        ("现有产品线", "Writing instruments / Drinkware / Tech / Bags / Apparel / Awards / Employee recognition / Mouse pads"),
    ]:
        kv(ws, k, v, r); r+=1
    section(ws, "【一轮判断】", r); r+=1
    for k, v in [
        ("综合优先级", "⭐⭐⭐⭐"),
        ("当前阶段", "待接触"),
        ("Fit Score", "69/100 → Priority B"),
        ("Value Score", "52/100 → Tier 3 Transactional"),
        ("开发投入级别", "Active（个性化LinkedIn + email sequence，qualification后报价）"),
        ("核心判断", "公司定位与Wischos正面重叠；20+年独立经营，决策链极短；ASI会员接受国际供应商；mission『长期关系』与Wischos material-led定位契合。"),
        ("ASI竞争分析", "ASI目录=北美中间商转售公模量产品。Wischos切入点：custom shape/specific material/gift set brief等ASI无法满足的需求，作为她的direct China metal partner。"),
        ("首轮限制", "不争ASI标准目录款；主攻custom/premium brief；WP-205金属书签（custom shape）是差异化最强切入点"),
        ("升级标准", "确认年订单≥200 pcs或有固定recurring客户项目 → 升⭐⭐⭐⭐⭐"),
        ("降级标准", "确认Canadian-made only或过去12个月无实质业务 → 降⭐⭐⭐"),
    ]:
        kv(ws, k, v, r); r+=1
    section(ws, "【供应商背景】", r); r+=1
    for k, v in [
        ("现有供应商模式", "ASI供应商网络转售，非自有制造能力；个性化通过ASI供应商协调，非自有设备"),
        ("社交媒体", "官网无社媒链接；LinkedIn个人页已找到，近期帖子无法抓取"),
        ("信息来源", "官网直接抓取 + LinkedIn搜索（2026-05-14）"),
    ]:
        kv(ws, k, v, r); r+=1
    set_col_widths(ws, [22, 70])

    # ── Sheet 2 ───────────────────────────────────────────────────────────────
    ws = wb.create_sheet("2-联系人")
    hdr(ws, "联系人档案")
    r = 2
    th(ws, ["姓名","职位","决策权","经验背景","LinkedIn","Email","电话","首选沟通渠道","备注"], r); r+=1
    row_data(ws, [
        "Lisa Janssen","Owner",
        "供应商选择/产品引入/定价/项目合作——完整决策权",
        "行业20+年；2003年起独立经营",
        "https://www.linkedin.com/in/lisa-janssen-1b63411/",
        "lisa@greatgiftideas.ca（官网确认）",
        "(647) 224-2399",
        "LinkedIn（首触）→ Email",
        "近期LinkedIn帖子无法抓取；无其他决策级联系人需查找"
    ], r)
    set_col_widths(ws, [18,12,35,28,42,28,16,20,40])
    ws.row_dimensions[3].height = 55
    for c in ws[3]:
        c.alignment = Alignment(wrap_text=True)

    # ── Sheet 3 ───────────────────────────────────────────────────────────────
    ws = wb.create_sheet("3-切入角度")
    hdr(ws, "三个切入角度（按优先级排序）")
    r = 2
    th(ws, ["排序","角度","为什么有效","话术 Hook","风险/注意"], r); r+=1
    angles = [
        ("⭐⭐⭐⭐⭐","Premium metal upgrade（最强）",
         "Lisa现有ASI目录有普通笔/饮品，但都是北美中间商转售的公模量产品。Wischos提供brass/titanium同类目更高材质版本，让她给客户的提案更差异化，利润空间更大。",
         "\"Your clients want gifts that stay on desks, not end up in a drawer. Brass pens and titanium tumblers have the kind of weight and finish that does.\"",
         "MOQ 100对boutique可能有阻力；提前准备50件pilot方案"),
        ("⭐⭐⭐⭐","Employee recognition blanks",
         "官网有明确employee recognition类目；ASI目录在premium metal选项薄弱；WP-205金属书签可做custom shape，recognition场景差异化最强。",
         "\"For milestone gifts, a custom metal piece engraved with a name or date carries different weight than a mug — and it's the kind of thing people keep.\"",
         "需确认她客户是否有自己的雕刻/个性化合作方"),
        ("⭐⭐⭐","Direct China metal partner（长期）",
         "ASI目录无法做custom-shaped金属件/特殊材质组合/独立brief开发。每次客户有这类需求，她需自己找直连供应商。Wischos填这个空缺。",
         "\"If a client brief calls for something in metal that the catalogue can't do — custom shape, specific material, a gift set idea — I'd be glad to help develop options.\"",
         "信任建立需时间；取决于她客户提出metal custom brief的频率"),
    ]
    for a in angles:
        row_data(ws, a, r); r+=1
    set_col_widths(ws, [8,25,42,55,30])

    # ── Sheet 4 ───────────────────────────────────────────────────────────────
    ws = wb.create_sheet("4-推荐产品配对")
    hdr(ws, "场景 × Wischos 产品配对")
    ws.cell(row=2, column=1, value="Top 4 首推聚焦：WP-101 / WP-205 / WP-401 / WP-103")
    ws["A2"].font = Font(bold=True, color="2F5597")
    r = 3
    th(ws, ["场景","推荐 SKU","产品名","为什么适合","Lisa侧卖法/客户场景/终端价"], r); r+=1
    products = [
        ("企业年终客户礼品","WP-101","Brass Crown Bolt-Action Pen",
         "高质感黄铜组合；桌面常用；logo雕刻持久；比ASI普通笔差异化",
         "精品礼品盒，$80–120/套终端价"),
        ("员工认可/里程碑（差异化最强）","WP-205","Custom Metal Bookmark",
         "可做custom shape（公司logo外形）；名字雕刻；独特性最强；ASI完全无法做",
         "Recognition场景，$40–60/件"),
        ("品牌饮品升级","WP-401","Pure Titanium Bottle",
         "纯钛材质独特；是ASI目录无法提供的差异化",
         "高端企业礼品，$80–120/件"),
        ("品牌笔类替换","WP-103","Tactical Stainless Steel Pen",
         "不锈钢+黄铜；重量感明显；比ASI普通笔差异化",
         "批量品牌笔，$25–35/件"),
        ("员工认可/里程碑","WP-206","Brass Spinning Top",
         "高感知价值；办公桌互动；适合10/20年纪念",
         "配礼盒，$50–80/件"),
        ("桌面件/recognition","WP-203","Brass Letter Opener",
         "高端感；桌面recognition场景",
         "Recognition+桌面礼品"),
        ("活动/峰会VIP礼品","WP-302","Brass Key Organizer",
         "EDC定位；实用+有材质感；适合科技/金融客户",
         "活动礼袋，$30–50/件"),
        ("VIP套件呈现","WGS-005 / WGS-006 / WGS-008","套件系列",
         "多件组合，呈现感强，作为场景展示（非首推）",
         "$120–200/套"),
    ]
    for p in products:
        row_data(ws, p, r); r+=1
    set_col_widths(ws, [28,12,28,40,35])

    # ── Sheet 5 ───────────────────────────────────────────────────────────────
    ws = wb.create_sheet("5-跟进时间表")
    hdr(ws, "多触点跟进策略（Active level）")
    r = 2
    th(ws, ["Day","渠道","动作","话术/物料","状态","实际日期/备注"], r); r+=1
    timeline = [
        ("Day 0","LinkedIn","发邀请（223 chars，待发）",
         "Impressive to see 20+ years building client relationships through branded promo. I help distributors build custom metal gift options from China - brass pens, titanium EDC and desk pieces. Thought it may be worth connecting.",
         "⬜ 待执行",""),
        ("Day 3","Google Maps","搜Great Gift Ideas → 若有Google Business，写同行真诚评论",
         "—","⬜ 待执行","若无Google Business则跳过"),
        ("Day 7","官网表单 / Email","LinkedIn未通过 → 官网表单触达（直达Lisa邮箱）\nLinkedIn已通过 → 发Email 1",
         "Email subject: Custom brass and titanium pieces for your corporate gifting clients",
         "⬜ 待触发",""),
        ("Day 14","Email","Email 2 + mockup图（把她logo P到WP-205+WP-101上）",
         "—","⬜ 待触发",""),
        ("Day 30","CRM","仍无回复 → 暂停，记入CRM","—","⬜ 待触发",""),
        ("Day 60","LinkedIn/Email","换角度2（recognition blanks）重新切入","—","⬜ 待触发",""),
        ("Day 90","LinkedIn","仍未通过 → 换hook（找到新业务动态后再试）","—","⬜ 待触发",""),
    ]
    for t in timeline:
        row_data(ws, t, r); r+=1
    set_col_widths(ws, [10,16,38,60,14,28])

    # ── Sheet 6 ───────────────────────────────────────────────────────────────
    ws = wb.create_sheet("6-话术库")
    hdr(ws, "话术库（LinkedIn / Email / 异议预案）")
    r = 2
    th(ws, ["类型","内容","备注/设计理由"], r); r+=1
    scripts = [
        ("LinkedIn邀请（⬜ 待发送，223 chars）",
         "Impressive to see 20+ years building client relationships through branded promo. I help distributors build custom metal gift options from China - brass pens, titanium EDC and desk pieces. Thought it may be worth connecting.",
         "主版本；≤300 chars；hook = 20+年经验"),
        ("LinkedIn邀请（备选，262 chars）",
         "Impressive to see 20+ years focused on recognition and corporate gifting — products that build lasting client relationships. I help distributors build custom metal gift options from China - brass, titanium and stainless pieces. Thought it may be worth connecting.",
         "备选；更强情绪价值"),
        ("通过后首条私信",
         "Hi Lisa, thanks for connecting. I'm [姓名].\n\nJust to give a clearer picture of what I do: I work mainly with metal gift projects - brass pens, titanium EDC, stainless desk pieces and drinkware for corporate gifting, employee recognition and branded promo briefs.\n\nI'm not limited to standard catalogue items; pieces can be developed around a client brief, material, finish or gift-set idea. The metal bookmark is one example - can be cut to a custom shape with logo, works well as a recognition or VIP gift item.\n\nI'll occasionally share relevant metal gift ideas. If a client brief involves a custom metal piece, happy to help explore options.\n\nThanks,\n[姓名]",
         "不重复邀请措辞；突出custom shape差异化；不发目录"),
        ("Email Subject（主选）",
         "Custom brass and titanium pieces for your corporate gifting clients",
         ""),
        ("Email Subject（备选）",
         "Metal gift options outside the ASI catalogue — for recognition and VIP briefs",
         ""),
        ("Email 1正文",
         "Hi Lisa,\n\nNoticed Great Gift Ideas has been building long-term client relationships through branded promo for over 20 years — and that your recognition and corporate gift programs are a big part of what you do.\n\nI work with distributors to source and develop custom metal gift options direct from China — a few directions that tend to work well for corporate gifting:\n- Brass bolt-action pens and stainless tactical pens (weight and finish that outlast standard catalogue pens)\n- Custom-shaped metal bookmarks (logo cut to shape — strong for recognition and milestone gifts)\n- Pure titanium tumblers and stainless desk pieces (categories where ASI options rarely go premium)\n\nFor briefs where the catalogue can't do custom shapes, specific materials, or a gift set idea, I develop options directly — no extra layer in between.\n\nHappy to share a few product references, or put together a short visual if any of these directions fit a current client program.\n\n[姓名]",
         "研究型开场；三个方向；不发目录；邀请下一步"),
        ("异议：已有供应商",
         "Not asking to replace your ASI suppliers — I work specifically on briefs where the catalogue can't do custom shapes or premium materials. Happy to be a backup for those.",
         ""),
        ("异议：MOQ太高",
         "I can start with a 50-piece pilot on one SKU so you can test with a client first.",
         "最低不低于30件"),
        ("异议：海外交期太长",
         "Air freight runs 7–10 days door-to-door. For repeat orders I can also arrange quicker turnaround from bonded stock.",
         ""),
    ]
    for s in scripts:
        row_data(ws, s, r); r+=1
    set_col_widths(ws, [30,80,35])

    # ── Sheet 7 ───────────────────────────────────────────────────────────────
    ws = wb.create_sheet("7-Sample-Pack")
    hdr(ws, "Sample Pack（Great Gift Ideas）")
    ws.cell(row=2, column=1, value="策略：免费寄样 + 自付运费（CAD 50–80）；连接通过并有兴趣后准备。")
    r = 3
    th(ws, ["#","SKU","产品","定制内容","用途/演示"], r); r+=1
    samples = [
        (1,"WP-205","Custom Metal Bookmark（带logo雕刻样，custom shape）","custom shape + logo雕刻","证明定制能力；recognition场景差异化最强——ASI完全无法做"),
        (2,"WP-401","Pure Titanium Bottle","logo engraving","高感知价值；让她感受材质；ASI无此款"),
        (3,"WP-101","Brass Crown Bolt-Action Pen","logo/name engraving","标准商务场景演示；重量感对比普通笔"),
        (4,"WP-403","Weighted Tumbler（如太重则只发规格单）","—（规格单）","饮品场景展示"),
        (5,"WP-203","Brass Letter Opener","logo engraving","桌面/recognition场景；高端感"),
        (6,"WP-103","Tactical SS Pen","logo engraving","男性化风格；补充场景"),
        ("A","Visual Brief","Photoshop/Canva","把她客户的logo P到WP-205+WP-101上","1张mockup图；展示成品效果"),
        ("B","Reference Card","1页规格单","—","产品规格 + 起订量（100/50 pilot）+ FOB价格区间参考"),
    ]
    for s in samples:
        row_data(ws, s, r); r+=1
    set_col_widths(ws, [4,18,35,25,48])

    # ── Sheet 8 ───────────────────────────────────────────────────────────────
    ws = wb.create_sheet("8-行动清单")
    hdr(ws, "本周可执行清单 + 长期扩展")
    r = 2
    th(ws, ["#","任务","类型","状态","完成日期/备注"], r); r+=1
    actions = [
        (1,"二轮客户调研完成（见MD文件）","内部研究","✅ 已完成","2026-05-14"),
        (2,"建立客户文件夹 + 创建本Excel","内部整理","✅ 已完成","2026-05-14"),
        (3,"同步写入总CRM：C039","内部整理","⬜ 待执行",""),
        (4,"发LinkedIn邀请（223 chars，已写好）","Day 0","⬜ 待执行",""),
        (5,"Google搜Great Gift Ideas → 写同行好评（若有Google Business）","Day 3","⬜ 待执行",""),
        (6,"LinkedIn未通过Day 7 → 官网表单触达","Day 7","⬜ 待触发",""),
        (7,"LinkedIn已通过 → 发Email 1","Day 7","⬜ 待触发",""),
        (8,"Day 14 → Email 2 + mockup图（WP-205+WP-101 P上客户logo）","Day 14","⬜ 待触发",""),
        (9,"准备Sample Pack（6件+reference card）","物料","⬜ 待执行",""),
        (10,"Canva/Photoshop做logo mockup（WP-205 custom shape效果图）","物料","⬜ 待执行",""),
        (11,"Day 30无回复 → 暂停，CRM标注【搁置】","Day 30","⬜ 待触发",""),
        (12,"Day 60 → 角度2（recognition blanks）重新切入","Day 60","⬜ 待触发",""),
        (13,"Day 90 LinkedIn仍未通过 → 查新动态，换hook再试","Day 90","⬜ 待触发",""),
        (14,"首单成后 → 探索recognition年度合作+转介绍","长期","⬜ 长期",""),
    ]
    for a in actions:
        row_data(ws, a, r); r+=1
    set_col_widths(ws, [4,55,18,14,28])

    # ── Sheet 9 ───────────────────────────────────────────────────────────────
    ws = wb.create_sheet("9-备注与扩展")
    hdr(ws, "长期扩展空间 + 关键备注")
    r = 2
    section(ws, "【长期扩展】", r); r+=1
    notes = [
        ("首单成后","推进recognition年度项目（员工周年/季度奖励），形成recurring订单"),
        ("转介绍路径","合作稳定后探索Lisa是否有ASI网络内其他distributor朋友可转介绍"),
        ("Private label","长期可开发private label系列（Lisa自有品牌metal gift line）"),
    ]
    for k, v in notes:
        kv(ws, k, v, r); r+=1
    section(ws, "【关键提醒（专属）】", r); r+=1
    reminders = [
        ("ASI竞争策略","不争ASI标准目录款；定位为ASI目录之外的direct China metal partner"),
        ("MOQ处理","首单MOQ 100有阻力时，主动提50件pilot；不低于30件"),
        ("身份表述","不冒充工厂；用'I source'/'I work with manufacturers'/'I help develop'"),
        ("差异化最强切入点","WP-205 custom shape是ASI完全无法做到的——始终在前期沟通中提到"),
        ("ASI vs Wischos","ASI目录=公模量产/北美中间商加价/无custom shape；Wischos=直连工厂/更灵活/特殊材质可做"),
        ("体量预估","估计5–15单/年，100–300 pcs/单；boutique规模有上限，但首单后有复购结构"),
    ]
    for k, v in reminders:
        kv(ws, k, v, r); r+=1
    section(ws, "【参考链接】", r); r+=1
    refs = [
        ("客户档案MD","临时/潜在/Great Gift Ideas - Lisa Janssen.md"),
        ("总CRM","临时/通用资料/Wischos_客户开发CRM.xlsx"),
        ("官网","greatgiftideas.ca"),
        ("LinkedIn","https://www.linkedin.com/in/lisa-janssen-1b63411/"),
    ]
    for k, v in refs:
        kv(ws, k, v, r); r+=1
    set_col_widths(ws, [22, 70])

    wb.save(path)
    print(f"✓ Saved: {path}")


# ══════════════════════════════════════════════════════════════════════════════
# UPDATE CRM
# ══════════════════════════════════════════════════════════════════════════════

def update_crm(crm_path):
    wb = openpyxl.load_workbook(crm_path)
    ws = wb["主板"]
    # Find next empty row after last data row
    last_row = ws.max_row
    while last_row > 1 and all(ws.cell(last_row, c).value is None for c in range(1, 13)):
        last_row -= 1
    next_row = last_row + 1

    # C038 – Branding Edge
    ws.cell(next_row, 1, "C038")
    ws.cell(next_row, 2, "Branding Edge")
    ws.cell(next_row, 3, "Graham Ellerman")
    ws.cell(next_row, 4, "小型本地促销品经销商 / 品牌定制装饰商 / 内部激光雕刻")
    ws.cell(next_row, 5, "加拿大 / Alberta")
    ws.cell(next_row, 6, "已发消息")
    ws.cell(next_row, 7, "B")
    ws.cell(next_row, 8, "2026-05-14")
    ws.cell(next_row, 9, "2026-06-14")
    ws.cell(next_row, 10, 1)
    ws.cell(next_row, 11, "★★★；LinkedIn邀请已发2026-05-14（版本B，无近期发帖）；激光雕刻切入角度；Light-touch，30天无回复暂停。")

    # C039 – Great Gift Ideas
    r2 = next_row + 1
    ws.cell(r2, 1, "C039")
    ws.cell(r2, 2, "Great Gift Ideas")
    ws.cell(r2, 3, "Lisa Janssen")
    ws.cell(r2, 4, "Boutique promo distributor / corporate gifting agency / ASI会员")
    ws.cell(r2, 5, "加拿大 / Ontario")
    ws.cell(r2, 6, "待接触")
    ws.cell(r2, 7, "B")
    ws.cell(r2, 8, "2026-05-14")
    ws.cell(r2, 9, "2026-05-21")
    ws.cell(r2, 10, 0)
    ws.cell(r2, 11, "⭐⭐⭐⭐；20+年经营；ASI会员；WP-205 custom shape是最强切入点；LinkedIn邀请待发。")

    wb.save(crm_path)
    print(f"✓ CRM updated: C038 + C039 added at rows {next_row}/{r2}")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    # 1. Create Canada subfolders
    be_dir = f"{BASE}/加拿大/Branding Edge - Graham Ellerman"
    gg_dir = f"{BASE}/加拿大/Great Gift Ideas - Lisa Janssen"
    os.makedirs(be_dir, exist_ok=True)
    os.makedirs(gg_dir, exist_ok=True)
    print(f"✓ Folders created")

    # 2. Move MD files
    be_md_src = f"{BASE}/潜在/Branding Edge - Graham Ellerman.md"
    gg_md_src = f"{BASE}/潜在/Great Gift Ideas - Lisa Janssen.md"
    be_md_dst = f"{be_dir}/Branding-Edge-Graham-Ellerman.md"
    gg_md_dst = f"{gg_dir}/Great-Gift-Ideas-Lisa-Janssen.md"

    if os.path.exists(be_md_src):
        shutil.move(be_md_src, be_md_dst)
        print(f"✓ Moved: {be_md_src} → {be_md_dst}")
    else:
        print(f"⚠ MD not found: {be_md_src}")

    if os.path.exists(gg_md_src):
        shutil.move(gg_md_src, gg_md_dst)
        print(f"✓ Moved: {gg_md_src} → {gg_md_dst}")
    else:
        print(f"⚠ MD not found: {gg_md_src}")

    # 3. Create Excel files
    make_branding_edge(f"{be_dir}/Branding-Edge-客户跟进.xlsx")
    make_great_gift_ideas(f"{gg_dir}/Great-Gift-Ideas-客户跟进.xlsx")

    # 4. Update CRM
    crm_path = f"{BASE}/通用资料/Wischos_客户开发CRM.xlsx"
    update_crm(crm_path)

    print("\n✅ All done.")
