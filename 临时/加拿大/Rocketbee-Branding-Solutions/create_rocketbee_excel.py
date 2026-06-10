#!/usr/bin/env python3
"""
Create Rocketbee Branding Solutions (C047) client Excel + update CRM.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = "/mnt/d/Project/Wischos trading website/临时"
CLIENT_DIR = f"{BASE}/加拿大/Rocketbee-Branding-Solutions"
CRM_PATH = f"{BASE}/通用资料/Wischos_客户开发CRM.xlsx"
CLIENT_ID = "C047"

# ── Helpers ──────────────────────────────────────────────────────────────────

def hdr(ws, text, row=1):
    ws.cell(row=row, column=1, value=text).font = Font(bold=True, size=12)

def subhdr(ws, text, row=2):
    ws.cell(row=row, column=1, value=text).font = Font(italic=True, color="595959")

def section(ws, label, row):
    c = ws.cell(row=row, column=1, value=label)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="595959")
    ws.merge_cells(f"A{row}:B{row}")

def kv(ws, key, value, row):
    ws.cell(row=row, column=1, value=key).font = Font(bold=True)
    ws.cell(row=row, column=2, value=value)

def th(ws, headers, row):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2F5597")

def row_data(ws, values, r):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.alignment = Alignment(wrap_text=True)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL: Rocketbee Branding Solutions – Mitch Rochette (C047)
# ══════════════════════════════════════════════════════════════════════════════

def make_rocketbee(path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: 客户档案 ─────────────────────────────────────────────────────
    ws = wb.create_sheet("1-客户档案")
    hdr(ws, "Rocketbee Branding Solutions — 客户档案")
    subhdr(ws, "建档日期：2026-05-19    来源：官网 / LinkedIn / ZoomInfo / GovShop / 联邦法人注册")
    r = 3
    kv(ws, "项目", "内容", r)
    ws["A3"].font = Font(bold=True); ws["B3"].font = Font(bold=True)
    r += 1
    section(ws, "【基础信息】", r); r += 1
    for k, v in [
        ("客户ID", CLIENT_ID),
        ("公司名", "Rocketbee Branding Solutions Inc."),
        ("联系人", "Mitch Rochette"),
        ("职位", "Owner / Director"),
        ("客户类型", "精品 promo distributor — B类（recognition & HR awards专项）"),
        ("地区", "Orleans (Ottawa), Ontario, Canada K4A 3W3"),
        ("网站", "https://www.rocketbee.ca"),
        ("邮箱", "mitch@rocketbee.ca（推断：ZoomInfo显示 m***@rocketbee.ca，中高可信，首次回复后确认）"),
        ("电话", "613-852-4811"),
        ("成立", "2012-12-18（联邦公司编号8382565）"),
        ("规模", "~3人（Mitch + Advertising Specialist Ayuning Tiyas + intern）"),
        ("营收估算", "~CAD $200K/年（ZoomInfo参考，非确认数字）"),
        ("行业协会", "PPPC 未确认（Ottawa地区大概率成员，无公开证据）"),
        ("GovShop", "已注册为政府供应商 ✓"),
        ("采购平台", "USB Promotions（rocketbee.usbpromotions.ca）— 现有供应商惯性来源"),
    ]:
        kv(ws, k, v, r); r += 1
    section(ws, "【一轮判断】", r); r += 1
    for k, v in [
        ("综合优先级", "⭐⭐⭐⭐ / B类"),
        ("当前阶段", "待接触"),
        ("Fit Score", "63/100 → Priority B"),
        ("Value Score", "55/100 → Tier 3 Transactional（边界Tier 2）"),
        ("开发投入级别", "Light-touch → Active（LinkedIn + 4-touch email；回复后升级）"),
        ("核心判断", "典型promo distributor，recognition/HR awards是其最高价值服务。USB Promotions catalog无多件金属套件，Wischos填补此空白。Mitch是Owner直接决策，25年+行业经验，Ottawa政府客户（Elections Canada）是战略价值点。"),
        ("关键发现", "USB Promotions平台依赖 = pitch必须是'catalog里没有的'；无直接China FOB进口经验；Q4 timing window现在已开启（5月）"),
        ("升级条件", "回复且对recognition/awards有兴趣 → 发分销商版Capability Sheet"),
        ("降级条件", "明确有固定中国供应商且不换 → 降C，2027 Q4再试"),
    ]:
        kv(ws, k, v, r); r += 1
    section(ws, "【供应商背景】", r); r += 1
    for k, v in [
        ("现有供应商", "USB Promotions（rocketbee.usbpromotions.ca）catalog平台为主；无直接中国工厂进口证据"),
        ("个性化技术", "协调外部供应商完成：debossing, embossing, foil stamping, laser etching, silkscreen/UV printing；自身无in-house设备"),
        ("已知产品线", "Recognition & HR awards / Promotional items / Branded apparel / Custom embroidery"),
        ("已知客户", "Elections Canada（联邦政府）/ First Nations Information Governance Centre / The Walker Group Ottawa"),
        ("信息来源", "官网 + LinkedIn + ZoomInfo + GovShop + 联邦法人注册 + BBB（2026-05-19）"),
    ]:
        kv(ws, k, v, r); r += 1
    set_col_widths(ws, [22, 72])

    # ── Sheet 2: 联系人 ───────────────────────────────────────────────────────
    ws = wb.create_sheet("2-联系人")
    hdr(ws, "联系人档案")
    r = 2
    th(ws, ["姓名", "职位", "决策权范围", "经验背景", "LinkedIn", "Email", "电话", "首选沟通渠道", "备注"], r)
    r += 1
    row_data(ws, [
        "Mitch Rochette",
        "Owner / Director",
        "供应商选择/产品引入/定价/项目合作——完整决策权（一人公司）",
        "25年+ promo行业；Adware Promotions（Owner）→ North Trek Promotions（Owner）→ VoX International Ltd.（Director Canada）→ Active Results Collaborative（Director of Ops）→ PromoMedia Group（Sales）→ Rocketbee 2012",
        "https://www.linkedin.com/in/mitch-rochette-4780934/",
        "mitch@rocketbee.ca（推断，待确认）",
        "613-852-4811",
        "LinkedIn（首触）→ Email",
        "唯一决策人；行业经验深厚，评估快；双语（英/法）可能；不喜欢泛泛推销，用具体业务细节切入"
    ], r)
    set_col_widths(ws, [18, 20, 30, 40, 44, 28, 16, 20, 40])
    ws.row_dimensions[3].height = 70
    for c in ws[3]:
        c.alignment = Alignment(wrap_text=True)

    # ── Sheet 3: 切入角度 ─────────────────────────────────────────────────────
    ws = wb.create_sheet("3-切入角度")
    hdr(ws, "三个切入角度（按优先级排序）")
    r = 2
    th(ws, ["排序", "角度", "为什么有效", "话术 Hook", "产品配对", "风险/注意"], r)
    r += 1
    angles = [
        ("⭐⭐⭐⭐⭐",
         "Recognition Gap — 定制金属套件填补catalog空白",
         "Rocketbee的recognition/HR awards是最高价值服务。USB Promotions等catalog供应商提供通用单品，无法做多件组合定制套件。WGS-005/006/008在标准加拿大promo catalog里无直接对应品，是结构性供应空白。",
         '"I noticed you handle recognition and HR award programs — wanted to share a curated metal gift set line that clients in similar markets are using for executive tenure gifts and onboarding kits, in case it fills a gap in what your current suppliers offer."',
         "WGS-005 Executive Set（高管tenure recognition）/ WGS-006 Onboarding Set（政府/NGO新员工）/ WGS-008 Quartet（会议代表premium礼品）",
         "Mitch可能希望看实物样品才推荐；无样品约束需用高质量产品图+FOB透明定价补偿"),
        ("⭐⭐⭐⭐",
         "Ottawa政府VIP Gifting — 特殊细分市场",
         "Rocketbee已服务Elections Canada和联邦NGO，GovShop注册确认政府供应商地位。Ottawa联邦政府的部长级礼品、外交礼品、峰会代表礼品是长期重复需求，对品质和重量感有明确要求。",
         '"Given your work with government clients in Ottawa, I thought you might see demand for premium branded metal sets for VIP gifting or delegation programs — something with more gravitas than a standard promo pen, but still fully customizable."',
         "WGS-005（高级官员桌面套件）/ 定制金属勋章（按需）/ WGS-008（代表团礼品套装）",
         "政府采购有Canadian content preference；初期不主动提；待建关系后评估"),
        ("⭐⭐⭐",
         "Eco-Friendly/Durable升级 — 替换塑料promo件",
         "Rocketbee官网和定位话语强调eco-friendly。行业趋势从一次性塑料promo转向耐久金属品。",
         '"I saw Rocketbee is focused on eco-friendly promotional solutions — we supply fully custom metal gift sets designed to last years, which is something clients increasingly want when they care about brand perception and sustainability."',
         "WP-401 Pure Titanium Bottle / WP-404 Bamboo-Groove SS Mug / WP-103 Tactical SS Pen",
         "最弱角度；Mitch见过很多eco-friendly pitch；需配具体产品图+定价才有说服力；作为跟进Angle 2备用"),
    ]
    for a in angles:
        row_data(ws, a, r); r += 1
    set_col_widths(ws, [8, 28, 42, 55, 35, 30])

    # ── Sheet 4: 推荐产品配对 ─────────────────────────────────────────────────
    ws = wb.create_sheet("4-推荐产品配对")
    hdr(ws, "场景 × Wischos 产品配对")
    ws.cell(row=2, column=1, value="Top 4 首推聚焦：WGS-005 / WGS-008 / WP-203 / WGS-006")
    ws["A2"].font = Font(bold=True, color="2F5597")
    r = 3
    th(ws, ["场景", "推荐 SKU", "产品名", "为什么适合", "Mitch侧卖法/终端场景/客单价"], r)
    r += 1
    products = [
        ("高管tenure recognition（5年/10年/退休）",
         "WGS-005 + WP-203 + WP-206",
         "Executive Set + Letter Opener + Brass Spinning Top",
         "高感知价值；多件组合比单品加价空间更大；可全套logo刻字",
         "CAD $120-200/套；premium recognition positioning"),
        ("政府/NGO新员工Onboarding kit",
         "WGS-006 + WP-208 + WP-101",
         "Onboarding Set + Device Stand + Brass Pen",
         "功能性强（桌面+书写）；批量需求稳定；政府对品质有要求",
         "CAD $80-130/套；批量稳定"),
        ("联邦政府VIP/外交礼品",
         "WGS-005 or WGS-008 + 定制金属勋章（按需）",
         "Executive/Quartet + Custom Medal",
         "Ottawa独特场景；单价高；protocol gift对重量感/精工要求高",
         "CAD $150-300+/件；单次10-50件"),
        ("会议/峰会代表礼品",
         "WGS-008 + WP-403",
         "The Quartet + Weighted Tumbler",
         "多件套有仪式感；可刻会议主题+logo；批量100-300件",
         "CAD $80-150/套"),
        ("年终企业礼品",
         "WGS-001 or WGS-002 + WP-206 + WP-403",
         "Gift Set + Spinning Top + Tumbler",
         "季节性高需求；可配Rocketbee礼盒定制包装",
         "CAD $60-120/套；批量100件+"),
        ("Eco-friendly promo升级",
         "WP-401 + WP-404 + WP-103",
         "Titanium Bottle + Bamboo-Groove Mug + Tactical Pen",
         "替换现有塑料单品；迎合客户可持续承诺；差异化定价",
         "CAD $40-80/件；批量200件+"),
    ]
    for p in products:
        row_data(ws, p, r); r += 1
    set_col_widths(ws, [28, 20, 30, 42, 35])

    # ── Sheet 5: 跟进时间表 ───────────────────────────────────────────────────
    ws = wb.create_sheet("5-跟进时间表")
    hdr(ws, "多触点跟进策略（Light-touch → Active）")
    r = 2
    th(ws, ["Day", "渠道", "动作", "话术/物料", "状态", "实际日期/备注"], r)
    r += 1
    timeline = [
        ("Day 0", "LinkedIn",
         "发送Mitch邀请（275 chars）",
         "Saw Rocketbee's recognition and HR awards programs. I like how you position recognition as something lasting. I help build custom metal gift options from China - brass, stainless and titanium pieces like desk sets, pens or award keepsakes. Thought it may be worth connecting.",
         "⬜ 待执行", ""),
        ("Day 0+2h", "Google Maps",
         "搜Rocketbee → 若有Google Business，写同行真诚好评一条",
         "—", "⬜ 待执行", "若无Google Business则跳过"),
        ("Day 3", "Instagram",
         "LinkedIn未通过 → 关注 @rocketbeebranding + 点赞最近3条动态",
         "—", "⬜ 待执行", ""),
        ("Day 3-5", "LinkedIn DM",
         "LinkedIn已通过 → 发首条私信",
         "见Sheet 6话术库",
         "⬜ 待触发", "连接通过后立即发"),
        ("Day 5", "官网Contact Form",
         "LinkedIn仍未通过 → 表单触达（小公司直达决策人邮箱）",
         "—", "⬜ 待触发", ""),
        ("Day 7", "Email",
         "发首封邮件 + 分销商版Capability Sheet附件",
         "Subject: Premium metal gift sets for recognition programs — filling the gap your catalog suppliers leave",
         "⬜ 待执行", "mitch@rocketbee.ca（待确认）"),
        ("Day 14", "LinkedIn/Email",
         "一次跟进，政府VIP角度",
         "见Sheet 6话术库Day 14跟进",
         "⬜ 待触发", ""),
        ("Day 30", "CRM",
         "暂停，归档；8月用Q4 timing hook重新切入",
         "—", "⬜ 待触发", ""),
        ("Day 60", "Email/LinkedIn",
         "Angle 3（eco-friendly升级）二次切入",
         "—", "⬜ 待触发", ""),
        ("Day 90", "CRM",
         "三次尝试前检查新trigger（LinkedIn新帖/展会/网站更新）",
         "—", "⬜ 待触发", ""),
    ]
    for t in timeline:
        row_data(ws, t, r); r += 1
    set_col_widths(ws, [12, 14, 32, 62, 14, 28])

    # ── Sheet 6: 话术库 ───────────────────────────────────────────────────────
    ws = wb.create_sheet("6-话术库")
    hdr(ws, "话术库（LinkedIn / DM / Email / 异议预案）")
    r = 2
    th(ws, ["类型", "内容", "备注/设计理由"], r)
    r += 1
    scripts = [
        ("LinkedIn邀请（⬜ 待发送，275 chars）",
         "Saw Rocketbee's recognition and HR awards programs. I like how you position recognition as something lasting. I help build custom metal gift options from China - brass, stainless and titanium pieces like desk sets, pens or award keepsakes. Thought it may be worth connecting.",
         "275 chars ✓；hook=recognition & HR awards（官网确认服务项）；认可=positioning recognition as something lasting；help build非supply；materials+开放例子；低压力结尾"),
        ("首条跟进私信（连接通过后Day 3-5）",
         "Hi Mitch, thanks for connecting. Just to give a clearer picture of what I do: I work mainly with metal gift projects — curated sets combining brass pens, desk accessories and EDC pieces — for recognition programs, executive gifting and onboarding kits.\n\nI'm not limited to catalogue items; pieces can be developed around a client brief, material or finish. The catalogue is just a starting point.\n\nI'll occasionally share relevant metal gift ideas — if a client brief involves a premium metal piece or gift set, I'd be happy to help explore options.\n\nThanks,\nTim\nWischos Gift | wischosgift.com",
         "不重复邀请措辞；定位为metal gift projects specialist；catalog只是参考；低压力后续；不发完整目录"),
        ("首封邮件 Subject",
         "Premium metal gift sets for recognition programs — filling the gap your catalog suppliers leave",
         "具体+有利益点；不泛泛；指向catalog gap痛点"),
        ("首封邮件 正文（Day 7）",
         "Hi Mitch,\n\nI'm reaching out because Rocketbee's recognition and HR awards work stood out to me — it's exactly the segment where I see the biggest gap between what catalog suppliers offer and what end clients actually want for executive-tier gifting.\n\nI'm behind Wischos Gift, a B2B exporter of fully custom metal gift sets sourced from specialist factories in Yangjiang, Dongguan, and Zhongshan. Our sets are curated combinations of metal pens, desk accessories, EDC tools, and drinkware — laser engraving, custom packaging, FOB pricing $18–$40/set at MOQ 100 sets.\n\nA few scenarios where your clients might find this useful:\n- Employee milestone and tenure recognition gifts (we have an Executive Desk Set and a four-piece \"Quartet\" that present well in branded gift boxes)\n- Onboarding kits for government or NGO clients who want something durable rather than plastic\n- VIP or delegation gifts for federal government events in Ottawa\n\nNone of our sets appear in standard Canadian promo catalogs — they're sourced and customized per order, so the recipient gets something that doesn't look like a catalog item.\n\nHappy to share our product deck with FOB pricing, lead times (25–35 days), and sample images. Would that be worth a look if you have a client project coming up this fall?\n\nBest,\nTim\nWischos Gift Trading\nwischosgift.com",
         "研究型开场；三个具体场景；不发目录；结尾低压力邀请；CASL合规（公开展示商业联系方式基础）"),
        ("Day 14跟进（政府VIP角度，≤3句）",
         "Hi Mitch — following up briefly in case this landed at a busy time. Given Rocketbee's federal government client base in Ottawa, I thought the angle of premium metal delegation or VIP gifts might be particularly relevant as clients scope fall recognition programs. Happy to share the product deck whenever it's convenient — no obligation.",
         "新角度（政府VIP）；低压力；不重复Day 7内容"),
        ("邮箱确认请求（LinkedIn私信通过后）",
         "Happy to send over a few specs — easiest by email. What's the best address?",
         "用于首条私信建立对话后，自然引出邮箱确认"),
        ("异议：已有供应商",
         "Not asking to replace your current suppliers — I fill the gap at the premium end, curated metal gift sets that catalog suppliers don't carry. Happy to be a backup option for specific projects.",
         "category expansion定位，非replacement"),
        ("异议：MOQ太高",
         "Standard MOQ is 100 sets. For a first project, happy to discuss a pilot run on one set — can you tell me more about the client brief?",
         "先问需求，再灵活讨论；不直接降到30以下"),
        ("异议：海外交期太长",
         "Production runs 25–35 days. Air freight to Ottawa is 7–10 days on top. Tell me your timeline and I'll confirm if it works.",
         "具体数字；先确认再承诺"),
        ("异议：只用本地供应",
         "Completely understood. I'll keep it in mind and check back if there's a project where direct sourcing makes sense.",
         "尊重，留门，CRM标注90天重查"),
    ]
    for s in scripts:
        row_data(ws, s, r); r += 1
    set_col_widths(ws, [32, 80, 38])

    # ── Sheet 7: Sample Pack ──────────────────────────────────────────────────
    ws = wb.create_sheet("7-Sample-Pack")
    hdr(ws, "Sample Pack（Rocketbee — 暂不寄样策略）")
    ws.cell(row=2, column=1,
            value="策略：Wischos无现货样品，不主动提样品。先用高质量产品图+spec sheet建立信任。对方明确有项目兴趣后，走自费采购流程（CAD 50-80运费）。首条私信不主动发catalog。")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 40
    r = 3
    th(ws, ["#", "内容", "形式", "触发条件", "备注"], r)
    r += 1
    samples = [
        (1, "WGS-005 Executive Set 产品图", "高质量JPG/PDF", "首封邮件或Day 14跟进时附", "主力展示套件；recognition场景最强"),
        (2, "WGS-008 The Quartet 产品图", "高质量JPG", "首封邮件附", "4件套呈现感；政府代表团场景"),
        (3, "WGS-006 Onboarding Set 产品图", "高质量JPG", "首封邮件附", "政府/NGO新员工onboarding"),
        (4, "Capability Sheet（分销商版）", "PDF附件", "首封邮件附", "分销商专版，含FOB价格区间"),
        ("A", "Mockup图（可选）", "Canva/Photoshop", "Day 14跟进时", "把Rocketbee/Elections Canada logo P到WGS-005上；杀手锏"),
        ("B", "实物样品", "按需采购", "对方确认有具体项目需求后", "自费采购，CAD 50-80运费；无现货库存"),
    ]
    for s in samples:
        row_data(ws, s, r); r += 1
    set_col_widths(ws, [4, 30, 20, 30, 40])

    # ── Sheet 8: 行动清单 ─────────────────────────────────────────────────────
    ws = wb.create_sheet("8-行动清单")
    hdr(ws, "行动清单（标准12项 + 长期扩展）")
    r = 2
    th(ws, ["#", "任务", "类型", "状态", "完成日期/备注"], r)
    r += 1
    actions = [
        (1, "双框架并行分析完成（见MD档案）", "内部研究", "✅ 已完成", "2026-05-19"),
        (2, "MD档案建立并归档", "内部整理", "✅ 已完成", "2026-05-19"),
        (3, "创建本Excel（9 sheet）", "内部整理", "✅ 已完成", "2026-05-19"),
        (4, "总CRM更新：C047 客户库+主板+跟进日志", "内部整理", "✅ 已完成（本脚本执行）", "2026-05-19"),
        (5, "发LinkedIn邀请（275 chars，recognition hook）", "Day 0", "⬜ 待执行", ""),
        (6, "Google搜Rocketbee → 写同行真诚好评（若有Google Business）", "Day 0+2h", "⬜ 待执行", ""),
        (7, "Instagram关注 @rocketbeebranding + 点赞最近3条动态", "Day 3", "⬜ 待执行", ""),
        (8, "LinkedIn通过后发首条私信（见Sheet 6）", "Day 3-5", "⬜ 待触发", ""),
        (9, "Day 7：发首封邮件 + 分销商版Capability Sheet", "Day 7", "⬜ 待触发", ""),
        (10, "Day 14：发跟进邮件（政府VIP角度，≤3句）", "Day 14", "⬜ 待触发", ""),
        (11, "Day 30无回复 → 暂停；8月用Q4 timing hook重触", "Day 30", "⬜ 待触发", ""),
        (12, "Day 60 → Angle 3（eco-friendly）二次切入", "Day 60", "⬜ 待触发", ""),
        (13, "Day 90 重查LinkedIn/网站触发信号，换hook再试", "Day 90", "⬜ 待触发", ""),
        (14, "首单成后：探索金属奖牌/lapel pins定制开发，进入Ottawa政府VIP渠道", "长期", "⬜ 长期", ""),
    ]
    for a in actions:
        row_data(ws, a, r); r += 1
    set_col_widths(ws, [4, 55, 18, 16, 28])

    # ── Sheet 9: 备注与扩展 ───────────────────────────────────────────────────
    ws = wb.create_sheet("9-备注与扩展")
    hdr(ws, "长期扩展空间 + 关键提醒 + 参考链接")
    r = 2
    section(ws, "【长期扩展】", r); r += 1
    for k, v in [
        ("首单成后", "引入定制金属奖牌/徽章（recognition program核心道具，Rocketbee现有catalog无此类别）"),
        ("Ottawa政府渠道", "如Rocketbee成功向政府客户推荐，可发展为政府protocol gifts固定供应商"),
        ("白标系列", "Rocketbee向客户提供'专属系列'金属礼品，Wischos代工"),
        ("转介绍", "合作稳定后探索Ottawa promo圈内转介绍"),
    ]:
        kv(ws, k, v, r); r += 1
    section(ws, "【关键提醒（专属）】", r); r += 1
    for k, v in [
        ("USB Promotions平台策略", "pitch必须是'catalog里没有的'，不要说'比现有供应商更好' — 定位是category expansion，非replacement"),
        ("无样品策略", "首次接触用高质量产品图+spec sheet；不主动提样品；如对方要样，走自费采购流程（CAD 50-80）"),
        ("身份表述", "不冒充工厂：用'I source'/'I work with manufacturers'/'I help develop'；不用'We are a factory'"),
        ("政府content preference", "初期不主动提产地限制问题；等建立关系后评估具体项目"),
        ("邮箱待确认", "LinkedIn通过后私信问：'Happy to send over a few specs — easiest by email. What's the best address?'"),
        ("MOQ策略", "标准100套；遇阻力先问需求；不轻易降到30以下；首单优先关系非volume"),
        ("CASL合规", "官网公开展示商业联系方式；冷邮件须含发件人身份+公司联系信息+退订机制"),
        ("体量预估", "年1-3单×100-300件；小而稳；promo distributor特点是复购稳定，长期价值高于看起来大的终端买家"),
    ]:
        kv(ws, k, v, r); r += 1
    section(ws, "【供应链位置】", r); r += 1
    ws.cell(r, 1, "中国工厂 → [USB Promotions / 北美catalog] → Rocketbee（distributor）→ 终端买家（政府/科技/NGO）→ 员工/代表")
    ws.merge_cells(f"A{r}:B{r}")
    ws.cell(r, 1).alignment = Alignment(wrap_text=True)
    r += 1
    ws.cell(r, 1, "Wischos切入点：填补catalog无法覆盖的高端定制多件套金属礼品")
    ws.cell(r, 1).font = Font(italic=True, color="2F5597")
    ws.merge_cells(f"A{r}:B{r}")
    r += 1
    section(ws, "【参考链接】", r); r += 1
    for k, v in [
        ("客户档案MD", "临时/加拿大/Rocketbee-Branding-Solutions/Rocketbee-Branding-Solutions-Mitch-Rochette.md"),
        ("总CRM", "临时/通用资料/Wischos_客户开发CRM.xlsx"),
        ("官网", "https://www.rocketbee.ca"),
        ("LinkedIn个人", "https://www.linkedin.com/in/mitch-rochette-4780934/"),
        ("LinkedIn公司", "https://ca.linkedin.com/company/rocketbee-branding-solutions-inc-"),
        ("GovShop", "https://govshop.com/supplier-profile/rocketbee-branding-solutions/"),
        ("联邦法人注册", "https://federalcorporation.ca/corporation/8382565"),
    ]:
        kv(ws, k, v, r); r += 1
    set_col_widths(ws, [24, 72])

    wb.save(path)
    print(f"✓ Saved: {path}")


# ══════════════════════════════════════════════════════════════════════════════
# UPDATE CRM
# ══════════════════════════════════════════════════════════════════════════════

def update_crm(crm_path):
    wb = openpyxl.load_workbook(crm_path)

    # ── 主板 ──────────────────────────────────────────────────────────────────
    ws_main = wb["主板"]
    last_row = ws_main.max_row
    while last_row > 1 and all(ws_main.cell(last_row, c).value is None for c in range(1, 13)):
        last_row -= 1
    nr = last_row + 1
    ws_main.cell(nr, 1, CLIENT_ID)
    ws_main.cell(nr, 2, f'=IFERROR(HYPERLINK("#客户库!A"&MATCH(A{nr},客户库!$A:$A,0),VLOOKUP(A{nr},客户库!$A:$B,2,0)),A{nr})')
    ws_main.cell(nr, 3, "Mitch Rochette")
    ws_main.cell(nr, 4, "精品 promo distributor / recognition & HR awards / 无自有中国供应链")
    ws_main.cell(nr, 5, "加拿大 / Ottawa, ON")
    ws_main.cell(nr, 6, "待接触")
    ws_main.cell(nr, 7, "B")
    ws_main.cell(nr, 8, "2026-05-19")
    ws_main.cell(nr, 9, "2026-05-19")
    ws_main.cell(nr, 10, f"=IFERROR(COUNTIF(跟进日志!$C:$C,A{nr}),0)")
    ws_main.cell(nr, 11, "⭐⭐⭐⭐；recognition/awards catalog gap切入；Q4 timing window开启；邮箱待确认；USB Promotions平台依赖；LinkedIn邀请待发（275 chars）。")
    print(f"✓ 主板 row {nr} added: {CLIENT_ID}")

    # ── 客户库 ────────────────────────────────────────────────────────────────
    ws_lib = wb["客户库"]
    last_lib = ws_lib.max_row
    while last_lib > 3 and all(ws_lib.cell(last_lib, c).value is None for c in range(1, 10)):
        last_lib -= 1
    lr = last_lib + 1
    ws_lib.cell(lr, 1, CLIENT_ID)
    ws_lib.cell(lr, 2, "Rocketbee Branding Solutions Inc.")
    ws_lib.cell(lr, 3, "Mitch Rochette")
    ws_lib.cell(lr, 4, "Owner / Director")
    ws_lib.cell(lr, 5, "精品 promo distributor — B类（recognition & HR awards专项）")
    ws_lib.cell(lr, 6, "https://www.linkedin.com/in/mitch-rochette-4780934/")
    ws_lib.cell(lr, 7, "mitch@rocketbee.ca（推断，待确认）")
    ws_lib.cell(lr, 8, "613-852-4811")
    ws_lib.cell(lr, 9, "—")
    ws_lib.cell(lr, 10, "Orleans (Ottawa), Ontario, Canada")
    ws_lib.cell(lr, 11, "~3人；~CAD $200K营收（估）；2012成立；GovShop注册")
    ws_lib.cell(lr, 12, "Recognition & HR awards / Promotional items / Branded apparel / Custom embroidery；USB Promotions catalog平台采购")
    ws_lib.cell(lr, 13, "2012")
    ws_lib.cell(lr, 14, "PPPC 未确认（Ottawa大概率成员）")
    print(f"✓ 客户库 row {lr} added: {CLIENT_ID}")

    # ── 跟进日志 ──────────────────────────────────────────────────────────────
    ws_log = wb["跟进日志"]
    last_log = ws_log.max_row
    while last_log > 2 and all(ws_log.cell(last_log, c).value is None for c in range(1, 9)):
        last_log -= 1
    xr = last_log + 1
    ws_log.cell(xr, 1, f"=IF(C{xr}=\"\",\"\",COUNTA($C$3:C{xr}))")
    ws_log.cell(xr, 2, "2026-05-19")
    ws_log.cell(xr, 3, CLIENT_ID)
    ws_log.cell(xr, 4, "Rocketbee Branding Solutions Inc.")
    ws_log.cell(xr, 5, "Mitch Rochette")
    ws_log.cell(xr, 6, "内部整理")
    ws_log.cell(xr, 7, "双框架并行分析 + 二轮开发流程")
    ws_log.cell(xr, 8, "双框架分析（一轮客户调查流程 + task-router×4 references）完成；MD档案建立；9-sheet Excel创建；总CRM更新。切入角度：recognition catalog gap（Angle 1最强）+ 政府VIP（Angle 2）+ eco-friendly升级（Angle 3）。LinkedIn邀请275 chars已写就，待发。")
    ws_log.cell(xr, 9, "已完成建档")
    print(f"✓ 跟进日志 row {xr} added: {CLIENT_ID}")

    wb.save(crm_path)
    print(f"✓ CRM saved: {crm_path}")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    excel_path = f"{CLIENT_DIR}/Rocketbee-Branding-Solutions-客户跟进.xlsx"
    make_rocketbee(excel_path)
    update_crm(CRM_PATH)
    print("\n✅ All done. Rocketbee C047 fully processed.")
